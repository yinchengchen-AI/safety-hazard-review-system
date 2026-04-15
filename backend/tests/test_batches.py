import io
import pytest
from httpx import AsyncClient


async def get_admin_token(client: AsyncClient) -> str:
    login = await client.post(
        "/api/v1/auth/login",
        data={"username": "admin", "password": "admin123"},
    )
    assert login.status_code == 200
    return login.json()["access_token"]


@pytest.mark.asyncio
async def test_preview_excel(client: AsyncClient):
    token = await get_admin_token(client)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["企业名称", "隐患描述", "隐患位置"])
    ws.append(["测试企业", "电线裸露", "车间A"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = await client.post(
        "/api/v1/batches/preview",
        files={"file": ("test.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 1
    assert len(data["items"]) == 1
    assert data["items"][0]["enterprise_name"] == "测试企业"
    assert data["items"][0]["errors"] == []
    assert "temp_token" in data


@pytest.mark.asyncio
async def test_preview_csv_with_empty_value(client: AsyncClient):
    token = await get_admin_token(client)
    csv_content = "企业名称,隐患描述,隐患位置\n,电线裸露,车间A\n"
    buffer = io.BytesIO(csv_content.encode("utf-8"))

    response = await client.post(
        "/api/v1/batches/preview",
        files={"file": ("test.csv", buffer, "text/csv")},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 1
    assert data["items"][0]["errors"] == ["企业名称不能为空"]


@pytest.mark.asyncio
async def test_import_and_duplicate_detection(client: AsyncClient):
    token = await get_admin_token(client)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["企业名称", "隐患描述", "隐患位置"])
    ws.append(["dup企业", "dup隐患", "dup位置"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # preview
    preview = await client.post(
        "/api/v1/batches/preview",
        files={"file": ("dup.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    temp_token = preview.json()["temp_token"]

    # first import
    import_res = await client.post(
        "/api/v1/batches/import",
        data={"temp_token": temp_token, "name": "dup_batch", "filename": "dup.xlsx"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert import_res.status_code == 200
    first = import_res.json()
    assert first["success_count"] == 1
    assert first["fail_count"] == 0

    # second import (duplicate)
    buffer2 = io.BytesIO()
    wb.save(buffer2)
    buffer2.seek(0)
    preview2 = await client.post(
        "/api/v1/batches/preview",
        files={"file": ("dup.xlsx", buffer2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    temp_token2 = preview2.json()["temp_token"]

    import_res2 = await client.post(
        "/api/v1/batches/import",
        data={"temp_token": temp_token2, "name": "dup_batch2", "filename": "dup.xlsx"},
        headers={"Authorization": f"Bearer {token}"},
    )
    assert import_res2.status_code == 200
    second = import_res2.json()
    assert second["success_count"] == 0
    assert second["fail_count"] == 1
    assert "重复数据（最近1个月内已存在）" in second["errors"][0]["reason"]


@pytest.mark.asyncio
async def test_template_download_excel(client: AsyncClient):
    token = await get_admin_token(client)
    response = await client.get(
        "/api/v1/batches/template?format=excel",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.asyncio
async def test_template_download_csv(client: AsyncClient):
    token = await get_admin_token(client)
    response = await client.get(
        "/api/v1/batches/template?format=csv",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert response.status_code == 200
    assert "text/csv" in response.headers["content-type"]


@pytest.mark.asyncio
async def test_list_batches_and_delete(client: AsyncClient):
    token = await get_admin_token(client)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["企业名称", "隐患描述", "隐患位置"])
    ws.append(["del企业", "del隐患", "del位置"])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    preview = await client.post(
        "/api/v1/batches/preview",
        files={"file": ("del.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": f"Bearer {token}"},
    )
    temp_token = preview.json()["temp_token"]

    import_res = await client.post(
        "/api/v1/batches/import",
        data={"temp_token": temp_token, "name": "del_batch", "filename": "del.xlsx"},
        headers={"Authorization": f"Bearer {token}"},
    )
    batch_id = import_res.json()["batch"]["id"]

    # list
    list_res = await client.get(
        "/api/v1/batches",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert list_res.status_code == 200
    batches = list_res.json()
    assert any(b["id"] == batch_id for b in batches)

    # delete
    del_res = await client.delete(
        f"/api/v1/batches/{batch_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert del_res.status_code == 204

    # list again
    list_res2 = await client.get(
        "/api/v1/batches",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert not any(b["id"] == batch_id for b in list_res2.json())

    # errors cleaned
    errors_res = await client.get(
        f"/api/v1/batches/{batch_id}/errors",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert errors_res.json() == []
